from configparser import ConfigParser
import logging
import openpyxl

def log_details():
    logging.basicConfig(
        filename="good.log",
        level=logging.INFO,
        format="%(asctime)s-%(levelname)s-%(message)s",
        datefmt="%Y-%m-%d %H:%M:%S %p"
    )
    return logging.getLogger()


logger = log_details()
logger.info("the test is started")


def config_details(key, value):
    conf = ConfigParser()
    conf.read("config.ini")
    return conf.get(key, value)


var = config_details("data", "city")
print(var)
print(config_details("data", "city"))
logger.info("config over")

logger.info("reading is started")

workbook=openpyxl.load_workbook("C:/Users/lenovo/PycharmProjects/Seleniumproject/data driven/files/test.xlsx")
sheet=workbook.active
rows=sheet.max_row
column=sheet.max_column
print(sheet.cell(1,4).value)
logger.info("good da ")
for r in range(1,rows+1):
    for c in range(1,column+1):
        print(sheet.cell(r,c).value, end="   ")
    print()