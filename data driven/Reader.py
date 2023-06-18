import openpyxl
def get_row(path,sheetname):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetname]
    return sheet.max_row

def get_columnn(path,sheetname):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetname]
    return sheet.max_column

def get_cell(path,sheetname,row_number,column_number):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetname]
    return (sheet.cell(row=row_number,column=column_number).value)

def get_cell_data(path,sheetname,row_number,column_number,data):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetname]
    sheet.cell(row=row_number, column=column_number).value=data
    workbook.save(path)

